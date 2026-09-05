"""Módulo para exportação de resultados em Excel e relatórios completos."""

from __future__ import annotations

import json
import logging
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from ..analysis.reports import ReportGenerator
from ..analysis.visualizations import ReviewVisualizer
from ..config import load_config
from ..db import read_papers
from ..processing.dedup import (
    audit_duplicate_candidates,
    build_identity_audit_rows,
    deterministic_identity_duplicate_mask,
    find_duplicates,
    normalize_doi,
)
from ..validation.derived_assets import sync_derived_assets

logger = logging.getLogger(__name__)


def _get_historical_dedup_stats() -> Tuple[int, int]:
    """Busca estatísticas históricas de deduplicação da tabela searches.

    PRISMA 2020 exige que 'Identification' mostre o total ORIGINAL coletado
    (antes de qualquer deduplicação). Como o banco atual já está limpo,
    precisamos buscar o initial_count registrado durante a coleta.

    Returns:
        Tuple[initial_count, total_removed]: Total original coletado e duplicatas removidas.
        Se não encontrado, retorna (0, 0).
    """
    try:
        config = load_config()
        db_path = Path(config.database.db_path)

        if not db_path.exists():
            return (0, 0)

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # Buscar o registro mais recente com dedup_stats
        cursor.execute("""
            SELECT results_summary FROM searches
            WHERE results_summary IS NOT NULL
            ORDER BY id DESC LIMIT 1
        """)

        row = cursor.fetchone()
        conn.close()

        if row and row[0]:
            try:
                data = json.loads(row[0])
                dedup_stats = data.get('dedup_stats', {})
                initial_count = dedup_stats.get('initial_count', 0)
                total_removed = dedup_stats.get('total_removed', 0)

                if initial_count > 0:
                    logger.info(
                        f"Histórico de dedup encontrado: initial_count={initial_count}, "
                        f"total_removed={total_removed}"
                    )
                    return (initial_count, total_removed)
            except (json.JSONDecodeError, TypeError):
                pass

        return (0, 0)

    except Exception as e:
        logger.warning(f"Erro ao buscar histórico de dedup: {e}")
        return (0, 0)


def to_excel(
    df: pd.DataFrame,
    output_path: Optional[Path] = None,
    sheet_name: str = "Papers",
    include_timestamp: bool = True,
    auto_adjust_columns: bool = True
) -> Path:
    """Exporta DataFrame para arquivo Excel formatado.

    Args:
        df: DataFrame com os dados
        output_path: Caminho de saída (se None, gera automaticamente)
        sheet_name: Nome da planilha
        include_timestamp: Se deve incluir timestamp no nome do arquivo
        auto_adjust_columns: Se deve ajustar largura das colunas

    Returns:
        Path do arquivo criado
    """
    if df.empty:
        logger.warning("DataFrame is empty, creating minimal Excel file")

    # Carregar configuração
    config = load_config()

    # Gerar nome do arquivo se necessário
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S") if include_timestamp else ""
        filename = f"systematic_review_{timestamp}.xlsx" if timestamp else "systematic_review.xlsx"
        output_path = Path(config.database.exports_dir) / filename

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Criar writer do pandas
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Escrever DataFrame principal
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Adicionar estatísticas em outra aba
        stats_df = create_statistics_dataframe(df)
        stats_df.to_excel(writer, sheet_name="Statistics", index=False)

        # Formatar as planilhas
        workbook = writer.book
        format_worksheet(workbook[sheet_name], auto_adjust_columns)
        format_worksheet(workbook["Statistics"], auto_adjust_columns)

    logger.info(f"Exported {len(df)} papers to {output_path}")
    return output_path


def to_excel_with_filters(
    df: pd.DataFrame,
    output_path: Optional[Path] = None,
    filters: Optional[Dict[str, List]] = None
) -> Path:
    """Exporta DataFrame para Excel com múltiplas abas baseadas em filtros.

    Args:
        df: DataFrame com os dados
        output_path: Caminho de saída
        filters: Dicionário com nome da aba e lista de condições

    Returns:
        Path do arquivo criado
    """
    config = load_config()

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path(config.database.exports_dir) / f"filtered_review_{timestamp}.xlsx"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Aba com todos os dados
        df.to_excel(writer, sheet_name="All Papers", index=False)

        # Abas filtradas
        if filters:
            for sheet_name, conditions in filters.items():
                filtered_df = df.copy()
                for condition in conditions:
                    filtered_df = filtered_df.query(condition)
                filtered_df.to_excel(writer, sheet_name=sheet_name[:31], index=False)  # Excel limit: 31 chars

        # Estatísticas
        stats_df = create_statistics_dataframe(df)
        stats_df.to_excel(writer, sheet_name="Statistics", index=False)

        # Formatar todas as planilhas
        workbook = writer.book
        for sheet in workbook.worksheets:
            format_worksheet(sheet, True)

    logger.info(f"Exported filtered results to {output_path}")
    return output_path


def create_statistics_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Cria DataFrame com estatísticas dos dados.

    Args:
        df: DataFrame com os papers

    Returns:
        DataFrame com estatísticas
    """
    stats = []

    # Estatísticas gerais
    stats.append({"Metric": "Total Papers", "Value": len(df)})

    if not df.empty:
        # Por ano
        if "year" in df.columns:
            year_counts = df["year"].value_counts().sort_index()
            stats.append({"Metric": "Papers per Year", "Value": ""})
            for year, count in year_counts.items():
                stats.append({"Metric": f"  {year}", "Value": count})

        # Por fonte
        if "source" in df.columns:
            source_counts = df["source"].value_counts()
            stats.append({"Metric": "Papers per Source", "Value": ""})
            for source, count in source_counts.items():
                stats.append({"Metric": f"  {source}", "Value": count})

        # Por search engine
        if "search_engine" in df.columns:
            engine_counts = df["search_engine"].value_counts()
            stats.append({"Metric": "Papers per Search Engine", "Value": ""})
            for engine, count in engine_counts.items():
                stats.append({"Metric": f"  {engine}", "Value": count})

        # Campos preenchidos
        stats.append({"Metric": "Field Completeness", "Value": ""})
        for col in df.columns:
            filled = df[col].notna().sum()
            percentage = (filled / len(df)) * 100
            stats.append({"Metric": f"  {col}", "Value": f"{filled} ({percentage:.1f}%)"})

        # Open Access
        if "is_open_access" in df.columns:
            open_access = df["is_open_access"].sum() if df["is_open_access"].dtype == bool else 0
            stats.append({"Metric": "Open Access Papers", "Value": f"{open_access} ({open_access/len(df)*100:.1f}%)"})

    return pd.DataFrame(stats)


def format_worksheet(worksheet, auto_adjust: bool = True):
    """Formata uma planilha do Excel.

    Args:
        worksheet: Planilha do openpyxl
        auto_adjust: Se deve ajustar largura das colunas
    """
    # Formatar cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Ajustar largura das colunas
    if auto_adjust:
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            worksheet.column_dimensions[column_letter].width = adjusted_width

    # Adicionar filtros
    if worksheet.max_row > 1:
        worksheet.auto_filter.ref = worksheet.dimensions


def export_for_analysis(
    df: pd.DataFrame,
    output_dir: Optional[Path] = None,
    formats: List[str] = ["xlsx", "csv", "json"]
) -> Dict[str, Path]:
    """Exporta dados em múltiplos formatos para análise.

    Args:
        df: DataFrame com os dados
        output_dir: Diretório de saída
        formats: Lista de formatos desejados

    Returns:
        Dicionário com formato e path do arquivo
    """
    config = load_config()

    if output_dir is None:
        output_dir = Path(config.database.exports_dir) / "analysis"

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    export_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    base_name = "papers"

    exported_files = {}

    if "xlsx" in formats:
        xlsx_path = output_dir / f"{base_name}.xlsx"
        df.to_excel(xlsx_path, index=False)
        exported_files["xlsx"] = xlsx_path
        logger.info(f"Exported to Excel: {xlsx_path}")

    if "csv" in formats:
        csv_path = output_dir / f"{base_name}.csv"
        # Strip only terminal whitespace from scalar text fields before the
        # line-oriented export. It is not bibliographic content and otherwise
        # makes the generated artifact fail ``git diff --check`` while making
        # no difference to the review data.
        df_csv = df.copy()
        for column in df_csv.select_dtypes(include=["object", "string"]).columns:
            df_csv[column] = df_csv[column].map(
                lambda value: "\n".join(
                    part.rstrip() for part in value.replace("\r", "").split("\n")
                )
                if isinstance(value, str)
                else value
            )
        df_csv.to_csv(
            csv_path,
            index=False,
            encoding='utf-8-sig',
            lineterminator='\n',
        )
        exported_files["csv"] = csv_path
        logger.info(f"Exported to CSV: {csv_path}")

    if "json" in formats:
        json_path = output_dir / f"{base_name}.json"
        # Converter int64 para int nativo antes de exportar
        df_json = df.copy()
        for col in df_json.select_dtypes(include=['int64']).columns:
            df_json[col] = df_json[col].astype(int)
        df_json.to_json(json_path, orient='records', indent=2, force_ascii=False)
        exported_files["json"] = json_path
        logger.info(f"Exported to JSON: {json_path}")

    return exported_files


def _select_best_duplicate(group: pd.DataFrame, preserve_stage: bool = True) -> pd.Series:
    """Seleciona o melhor registro entre duplicatas baseado em critérios de qualidade.

    Critérios (em ordem de prioridade):
    1. Registro com abstract mais longo
    2. Maior número de citações
    3. Com open_access_pdf disponível
    4. Primeiro registro (mais antigo na coleta)

    IMPORTANTE: Preserva selection_stage mais alto do original quando duplicate vencedor
    tiver estágio inferior (ex: duplicate com abstract melhor mas stage='screening' não
    substitui original com stage='included').

    Args:
        group: DataFrame com registros duplicados
        preserve_stage: Se True, propaga selection_stage mais alto do original

    Returns:
        Melhor registro do grupo com stage preservado se aplicável
    """
    if len(group) == 1:
        return group.iloc[0].copy()

    # Criar score de qualidade
    scores = pd.Series(0.0, index=group.index)

    # 1. Abstract length (peso 40%)
    if 'abstract' in group.columns:
        abstract_len = group['abstract'].fillna('').astype(str).str.len()
        if abstract_len.max() > 0:
            scores += (abstract_len / abstract_len.max()) * 40

    # 2. Citation count (peso 30%)
    if 'citation_count' in group.columns:
        citations = group['citation_count'].fillna(0).astype(float)
        if citations.max() > 0:
            scores += (citations / citations.max()) * 30

    # 3. Open access PDF (peso 20%)
    if 'open_access_pdf' in group.columns:
        has_pdf = group['open_access_pdf'].notna() & (group['open_access_pdf'] != '')
        scores += has_pdf.astype(int) * 20

    # 4. Ordem de coleta (peso 10% - favorece primeiro)
    scores += (len(group) - group.reset_index(drop=True).index) * (10.0 / len(group))

    # Retornar registro com maior score
    best_idx = scores.idxmax()
    best = group.loc[best_idx].copy()

    # CORREÇÃO: Preservar selection_stage mais alto do original
    if preserve_stage and 'selection_stage' in group.columns:
        # Encontrar registro original (não duplicata)
        originals = group[~group['is_duplicate'].astype(bool)]
        if not originals.empty:
            original = originals.iloc[0]

            # Hierarquia de prioridade de estágios
            stage_priority = {'included': 3, 'eligibility': 2, 'screening': 1}
            original_priority = stage_priority.get(original.get('selection_stage'), 0)
            best_priority = stage_priority.get(best.get('selection_stage'), 0)

            # Propagar stage do original se for mais alto
            if original_priority > best_priority:
                old_stage = best.get('selection_stage')
                best['selection_stage'] = original['selection_stage']
                if 'status' in original.index:
                    best['status'] = original['status']
                logger.info(
                    f"Stage preservado: Duplicate vencedor tinha stage '{old_stage}' "
                    f"mas original era '{original.get('selection_stage')}' - propagando stage mais alto"
                )

    return best


def _compute_prisma_stats_from_df(
    df: pd.DataFrame,
    unique_subset: Optional[pd.DataFrame] = None
) -> dict:
    """Compute PRISMA stats from DataFrame.

    PRISMA 2020 flow correto:
    - identification: Total de registros presentes no snapshot exportado
    - duplicates_removed: Registros removidos por identidade DOI/URL exata
    - screening: Registros sem duplicidade determinística disponíveis para triagem
    - screening_excluded: Registros excluídos NA TRIAGEM (selection_stage='screening')
    - eligibility: Registros que PASSARAM triagem (selection_stage='eligibility' ou 'included')
    - eligibility_excluded: Registros excluídos NA ELEGIBILIDADE (selection_stage='eligibility')
    - included: Registros finalmente incluídos (selection_stage='included')

    O histórico da tabela searches só é usado quando sua contagem é compatível
    com o DataFrame atual. Um registro histórico incompatível não pode ser
    misturado ao snapshot vigente, pois produziria um fluxo PRISMA impossível.

    Args:
        df: DataFrame com papers do snapshot (flags operacionais preservadas)
        unique_subset: Subconjunto único para consistência (opcional)

    Returns:
        Dicionário com estatísticas PRISMA
    """
    stats = {}

    raw_rows = int(len(df))

    # Normalizar DOIs para métricas de integridade e identidade determinística.
    if 'doi' in df.columns:
        normalized = (
            df['doi']
            .fillna('')
            .astype(str)
            .str.strip()
            .str.lower()
        )
        distinct_dois = {d for d in normalized if d}
    else:
        distinct_dois = set()

    distinct_count = int(len(distinct_dois)) if distinct_dois else raw_rows

    stats['raw_rows'] = raw_rows
    stats['distinct_doi'] = distinct_count

    # Subconjunto único para cálculos de estágios (define antes de usar)
    if unique_subset is not None:
        unique_df = unique_subset.copy()
    elif 'is_duplicate' in df.columns:
        duplicate_mask = deterministic_identity_duplicate_mask(df)
        unique_df = df[~duplicate_mask].copy()
    else:
        # DataFrames externos podem não carregar a flag persistida. Neles,
        # aplicar somente o fallback por DOI normalizado. Títulos repetidos
        # permanecem candidatos de auditoria, pois a igualdade textual não
        # prova que dois registros são a mesma publicação.
        fallback_duplicate = pd.Series(False, index=df.index)
        if 'doi' in df.columns:
            normalized_doi = df['doi'].fillna('').astype(str).map(normalize_doi)
            fallback_duplicate |= (
                normalized_doi.ne('')
                & normalized_doi.duplicated(keep='first')
            )
        unique_df = df[~fallback_duplicate].copy()

    # O histórico só pode ser reutilizado quando fecha aritmeticamente com o
    # snapshot atual. Caso contrário, trata-se de uma execução anterior ou de
    # uma coleta diferente e deve permanecer apenas como informação de log.
    historical_initial, historical_removed = _get_historical_dedup_stats()

    # Contagem de linhas sem flag operacional após a identificação
    screening_count = int(len(unique_df))

    historical_is_compatible = (
        historical_initial > 0
        and historical_removed >= 0
        and historical_initial - historical_removed == screening_count
        and historical_initial >= raw_rows
    )

    if historical_is_compatible:
        stats['identification'] = historical_initial

        computed_removed = max(0, historical_initial - screening_count)
        stats['duplicates_removed'] = computed_removed

        # Se o valor histórico não bate com o cálculo atual, registrar divergência
        if historical_removed and historical_removed != computed_removed:
            delta = computed_removed - historical_removed
            logger.warning(
                "PRISMA: divergência entre histórico e banco atual. "
                f"historical_removed={historical_removed}, computed_removed={computed_removed}, delta={delta}"
            )
        else:
            logger.info(
                f"PRISMA usando histórico: identification={historical_initial}, "
                f"duplicates_removed={computed_removed}"
            )
    else:
        # Usar a contagem atual garante que o fluxo seja aritmeticamente
        # coerente com os registros efetivamente exportados.
        stats['identification'] = raw_rows
        stats['duplicates_removed'] = max(0, stats['identification'] - screening_count)
        if historical_initial > 0:
            logger.warning(
                "Histórico de dedup incompatível com o snapshot atual; "
                f"não será usado no PRISMA (historical={historical_initial}, "
                f"removed={historical_removed}, current={raw_rows}, "
                f"screening={screening_count})."
            )

    stats['screening'] = screening_count

    # Cálculo baseado em selection_stage APENAS sobre registros únicos
    if not unique_df.empty and 'selection_stage' in unique_df.columns:
        stage_series = unique_df['selection_stage'].fillna('').astype(str).str.lower()
        if 'status' in unique_df.columns:
            status_series = unique_df['status'].fillna('').astype(str).str.lower()
        else:
            status_series = pd.Series('', index=unique_df.index)

        # Excluídos na triagem exigem status de exclusão
        screening_mask = stage_series == 'screening'
        screening_excluded = screening_mask & status_series.str.contains('exclu')
        stats['screening_excluded'] = int(screening_excluded.sum())

        # Passaram triagem: elegibilidade ou incluídos
        passed_screening = stage_series.isin(['eligibility', 'included'])
        stats['eligibility'] = int(passed_screening.sum())

        # Excluídos na elegibilidade
        eligibility_mask = stage_series == 'eligibility'
        eligibility_excluded = eligibility_mask & status_series.str.contains('exclu')
        stats['eligibility_excluded'] = int(eligibility_excluded.sum())

        # Incluídos finais
        included_mask = stage_series == 'included'
        stats['included'] = int(included_mask.sum())
    else:
        stats['screening_excluded'] = 0
        stats['eligibility'] = int(len(unique_df))
        stats['eligibility_excluded'] = 0
        stats['included'] = int(len(unique_df))

    # Preserve a flag persistida e a deduplicação determinística como
    # dimensões separadas. Títulos repetidos continuam sendo candidatos.
    stats['deduplication_audit'] = audit_duplicate_candidates(df)

    identification = int(stats.get('identification', raw_rows))
    eligibility = int(stats.get('eligibility', 0))
    stats['stage_percentages'] = {
        'screening_excluded_of_identification': round(
            100 * stats['screening_excluded'] / identification, 2
        ) if identification else 0.0,
        'screening_advanced_of_identification': round(
            100 * eligibility / identification, 2
        ) if identification else 0.0,
        'eligibility_excluded_of_eligibility': round(
            100 * stats['eligibility_excluded'] / eligibility, 2
        ) if eligibility else 0.0,
        'included_of_eligibility': round(
            100 * stats['included'] / eligibility, 2
        ) if eligibility else 0.0,
        'included_of_identification': round(
            100 * stats['included'] / identification, 2
        ) if identification else 0.0,
    }

    # AUDITORIA: Métricas de validação para preservação de stage
    if unique_subset is not None:
        raw_included = len(df[df['selection_stage'] == 'included'])
        included_without_operational_flag = stats['included']
        duplicates_included = len(df[
            (df['selection_stage'] == 'included') &
            (df['is_duplicate'].astype(bool))
        ])

        expected_delta = duplicates_included
        actual_delta = raw_included - included_without_operational_flag

        stats['_audit'] = {
            'raw_included_total': raw_included,
            'duplicates_marked_included': duplicates_included,
            'included_without_operational_flag': included_without_operational_flag,
            'expected_delta': expected_delta,
            'actual_delta': actual_delta
        }

        if actual_delta != expected_delta:
            logger.warning(
                f"⚠️ VALIDAÇÃO FALHOU: Esperava remover {expected_delta} papers 'included' duplicados, "
                f"mas removeu {actual_delta}. Delta incorreto: {actual_delta - expected_delta} papers perdidos."
            )
        else:
            logger.info(
                f"✓ Validação passou: Removidos {actual_delta} papers 'included' duplicados conforme esperado"
            )

    return stats


def get_best_duplicates(df: pd.DataFrame) -> pd.DataFrame:
    """Retorna DataFrame com os melhores registros entre duplicatas.

    Para cada grupo de duplicatas, seleciona o registro com melhor qualidade.
    Registros sem flag operacional (is_duplicate=False) são mantidos como estão;
    isso não equivale a unicidade bibliográfica.

    CORREÇÃO (2025-11-24): Agora remove originals quando um duplicate melhor é encontrado.

    Args:
        df: DataFrame com papers (incluindo duplicatas)

    Returns:
        DataFrame com melhores registros (duplicatas removidas, originals substituídos se necessário)
    """
    if df.empty:
        return df

    # Se não tem coluna is_duplicate, retorna como está
    if 'is_duplicate' not in df.columns:
        return df

    # Separar únicos e duplicatas.  Cada grupo processado abaixo substitui
    # todos os seus membros por exatamente um representante; isso evita que
    # o registro original seja mantido e o "melhor duplicado" seja anexado
    # novamente (o comportamento anterior podia deixar DOI repetido).
    unique_papers = df[~df['is_duplicate'].astype(bool)].copy()
    duplicate_papers = df[df['is_duplicate'].astype(bool)].copy()

    if duplicate_papers.empty:
        return unique_papers

    processed_unique_indices: set[Any] = set()
    processed_duplicate_indices: set[Any] = set()
    representatives: list[pd.Series] = []

    def _reference_matches(reference: Any, candidates: pd.DataFrame) -> pd.DataFrame:
        if not isinstance(reference, str):
            return candidates.iloc[0:0]
        if reference.startswith('DOI:') and 'doi' in candidates.columns:
            reference_doi = normalize_doi(reference[4:])
            normalized_doi = candidates['doi'].fillna('').astype(str).map(normalize_doi)
            return candidates[normalized_doi == reference_doi]
        if reference.startswith('URL:') and 'url' in candidates.columns:
            reference_url = reference[4:].strip().lower()
            normalized_url = candidates['url'].fillna('').astype(str).str.strip().str.lower()
            return candidates[normalized_url == reference_url]
        return candidates.iloc[0:0]

    for dup_index, dup_row in duplicate_papers.iterrows():
        if dup_index in processed_duplicate_indices:
            continue

        reference = dup_row.get('duplicate_of')
        if isinstance(reference, str) and reference:
            same_reference = duplicate_papers[
                duplicate_papers['duplicate_of'].fillna('').astype(str) == reference
            ]
            duplicate_group = same_reference
            original_group = _reference_matches(reference, unique_papers)
        else:
            duplicate_group = duplicate_papers.loc[[dup_index]]
            original_group = duplicate_papers.iloc[0:0]

        full_group = pd.concat([original_group, duplicate_group], ignore_index=True)
        best = _select_best_duplicate(full_group)

        # Select the best metadata record, but preserve the most advanced
        # selection stage represented by the identity group. The explicit
        # provenance field prevents the propagated stage from being mistaken
        # for the stage originally stored on the metadata-winning row.
        stage_source_id = None
        if 'selection_stage' in full_group.columns:
            stage_priority = {'screening': 1, 'eligibility': 2, 'included': 3}
            highest_stage_index = max(
                range(len(full_group)),
                key=lambda index: stage_priority.get(
                    str(full_group.iloc[index].get('selection_stage', '')).lower(), 0
                ),
            )
            highest_stage_row = full_group.iloc[highest_stage_index]
            stage_source_id = highest_stage_row.get('id', highest_stage_index)
            best['selection_stage'] = highest_stage_row.get('selection_stage')
            if 'status' in full_group.columns:
                best['status'] = highest_stage_row.get('status')
            if 'exclusion_reason' in full_group.columns:
                best['exclusion_reason'] = highest_stage_row.get('exclusion_reason')
        if stage_source_id is not None:
            best['stage_source_id'] = stage_source_id
        best['is_duplicate'] = False
        best['duplicate_of'] = None
        representatives.append(best)

        processed_unique_indices.update(original_group.index)
        processed_duplicate_indices.update(duplicate_group.index)

    remaining_unique = unique_papers.drop(index=list(processed_unique_indices), errors='ignore')
    result = pd.concat([remaining_unique, pd.DataFrame(representatives)], ignore_index=True)
    return result


def export_complete_review(
    df: pd.DataFrame,
    stats: Optional[Dict] = None,
    config: Optional[Dict] = None,
    output_dir: Optional[Path] = None,
    fulltext_stats: Optional[Dict] = None
) -> Dict[str, Path]:
    """Export complete review with Excel, visualizations and reports.

    Args:
        df: DataFrame with papers (incluindo duplicatas marcadas)
        stats: PRISMA statistics
        config: Configuration used
        output_dir: Output directory
        fulltext_stats: Full-text extraction statistics (optional)

    Returns:
        Dictionary with paths to generated files

    Note:
        Duplicatas (is_duplicate=True) são automaticamente filtradas.
        As análises, visualizações e relatórios usam apenas registros sem flag
        operacional; candidatos de identidade permanecem auditáveis.
    """
    cfg = load_config()

    if output_dir is None:
        output_dir = Path(cfg.database.exports_dir)

    output_dir = Path(output_dir)
    export_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    logger.info("Exporting complete review results...")

    exported_files = {}

    try:
        # Ensure we have data from DB if empty
        if df is None or df.empty:
            try:
                df = read_papers(cfg)
                logger.info("Reloaded papers from DB for export (canonical source)")
            except Exception as e:
                logger.warning(f"Failed to reload DB for export: {e}")
                df = pd.DataFrame()

        # Data already deduplicated by pipeline - just normalize DOIs for consistency
        if not df.empty and 'doi' in df.columns:
            df = df.copy()
            df['doi'] = df['doi'].fillna('').astype(str).apply(normalize_doi)

        raw_df = df

        # The operational DB currently contains 0 persisted duplicate flags,
        # but it still has repeated exact DOI/URL identities. Re-run the
        # deterministic identity pass at export time so the PRISMA flow does
        # not mistake an empty flag column for a duplicate-free snapshot.
        marked_df = find_duplicates(raw_df)

        # AUDITORIA: Capturar contagens de stage antes da deduplicação
        raw_stages = marked_df['selection_stage'].value_counts().to_dict() if 'selection_stage' in marked_df.columns else {}

        # Selecionar um representante por identidade determinística, mantendo
        # títulos apenas semelhantes fora da deduplicação automática.
        df_best = get_best_duplicates(marked_df) if 'is_duplicate' in marked_df.columns else marked_df
        if 'is_duplicate' in df_best.columns:
            # Após seleção de melhores duplicatas, garantir flag consistente
            df_best = df_best.copy()
            if 'is_duplicate' in df_best.columns:
                # Recalcular flag (todos agora únicos)
                df_best['is_duplicate'] = False
                df_best['duplicate_of'] = None
        df_for_export = df_best
        logger.info(f"Registros para export (após melhor duplicata): {len(df_for_export)}")

        # AUDITORIA: Comparar contagens de stage após deduplicação
        if raw_stages and 'selection_stage' in df_for_export.columns:
            export_stages = df_for_export['selection_stage'].value_counts().to_dict()
            logger.info(f"Contagem de stages antes dedup: {raw_stages}")
            logger.info(f"Contagem de stages após dedup: {export_stages}")

            # Avisos para perdas em stages críticos
            for stage in ['included', 'eligibility']:
                raw_count = raw_stages.get(stage, 0)
                export_count = export_stages.get(stage, 0)
                delta = raw_count - export_count

                if delta > 0:
                    logger.warning(
                        f"⚠️ Stage '{stage}' perdeu {delta} papers durante deduplicação "
                        f"({raw_count} -> {export_count})"
                    )
                elif delta < 0:
                    logger.error(
                        f"❌ Stage '{stage}' GANHOU papers durante deduplicação - problema de integridade! "
                        f"({raw_count} -> {export_count})"
                    )

        # Calcular estatísticas uma única vez (inclui raw_rows/distinct_doi)
        # Keep the audit anchored to the raw DB snapshot so the report can
        # distinguish persisted flags (possibly zero) from derived identity
        # removals performed for this export.
        local_stats = _compute_prisma_stats_from_df(raw_df, df_for_export)
        if stats:
            prisma_keys = {
                'identification', 'duplicates_removed', 'screening',
                'screening_excluded', 'eligibility', 'eligibility_excluded', 'included'
            }
            for k, v in stats.items():
                if k not in prisma_keys:
                    local_stats[k] = v

        # 1. Excel files (moved to analysis folder to avoid duplicate basenames)
        analysis_dir = output_dir / "analysis"
        analysis_dir.mkdir(parents=True, exist_ok=True)
        identity_audit_path = analysis_dir / "deduplication_identity_audit.csv"
        identity_audit_rows = build_identity_audit_rows(marked_df, retained_df=df_for_export)
        audit_df = pd.DataFrame(
            identity_audit_rows,
            columns=[
                "duplicate_id", "retained_id", "identifier_type", "identifier",
                "duplicate_title", "retained_title", "duplicate_stage",
                "retained_stage", "duplicate_status", "retained_status",
                "retained_stage_source_id", "decision",
            ],
        )
        for column in audit_df.select_dtypes(include=["object", "string"]).columns:
            audit_df[column] = audit_df[column].map(
                lambda value: "\n".join(
                    part.rstrip() for part in value.replace("\r", "").split("\n")
                )
                if isinstance(value, str)
                else value
            )
        audit_df.to_csv(
            identity_audit_path,
            index=False,
            encoding="utf-8-sig",
            lineterminator="\n",
        )
        exported_files["deduplication_audit"] = identity_audit_path

        excel_path = to_excel_with_filters(
            df_for_export,
            analysis_dir / "revisao_sistematica.xlsx"
        )
        exported_files["excel"] = excel_path

        # 2. Multi-format export (analysis artifacts) - usa dados sem flag
        # operacional; candidatos de identidade permanecem no snapshot.
        analysis_dir = output_dir / "analysis"
        analysis_files = export_for_analysis(
            df_for_export,
            analysis_dir,
            formats=["xlsx", "csv", "json"]
        )
        exported_files.update(analysis_files)

        # 3. Visualizations (usa o snapshot exportado + stats consistentes)
        visualizer = ReviewVisualizer(output_dir / "visualizations")
        chart_paths = visualizer.generate_all_visualizations(df_for_export, local_stats)
        exported_files["charts"] = chart_paths

        # The canonical export is the only source allowed to update the
        # downstream TCC and Slidev copies.  Custom output directories remain
        # useful for experiments and must not overwrite published artifacts.
        canonical_exports_dir = Path(__file__).resolve().parents[2] / "exports"
        if output_dir.resolve() == canonical_exports_dir.resolve():
            exported_files["derived_assets_manifest"] = sync_derived_assets()

        # 4. Reports (usa o snapshot exportado)
        report_generator = ReportGenerator(output_dir / "reports")

        # Summary report (passa DataFrame filtrado e stats originais + fulltext_stats)
        summary_path = report_generator.generate_summary_report(df_for_export, local_stats, config, fulltext_stats=fulltext_stats)
        exported_files["summary_report"] = summary_path

        # Papers report (included papers only - já filtrado) - passa fulltext_stats para incluir dados de extração
        if "selection_stage" in df_for_export.columns:
            papers_path = report_generator.generate_papers_report(df_for_export, "included", fulltext_stats=fulltext_stats)
            exported_files["papers_report"] = papers_path

        # Gap analysis
        gap_path = report_generator.generate_gap_analysis(df_for_export)
        exported_files["gap_analysis"] = gap_path

        # Optional cleanup: remove analysis-level artifacts after they have
        # been used to generate the final reports (keeps reports and visuals)
        # Default behavior: keep analysis artifacts unless explicitly requested
        # to cleanup. This preserves `papers.csv/json/xlsx` by default.
        cleanup_analysis = False if config is None else bool(config.get("cleanup_analysis", False))
        if cleanup_analysis:
            try:
                for ext in ("csv", "json", "xlsx"):
                    p = analysis_dir / f"papers.{ext}"
                    if p.exists():
                        p.unlink()
                        logger.info(f"Removed analysis artifact: {p}")
                # Also remove the separate revisao_sistematica.xlsx if present
                revisao = analysis_dir / "revisao_sistematica.xlsx"
                if revisao.exists():
                    revisao.unlink()
                    logger.info(f"Removed analysis artifact: {revisao}")
            except Exception as e:
                logger.warning(f"Failed to cleanup analysis artifacts: {e}")

        logger.info(f"Complete review exported to {output_dir}")
        logger.info(f"Generated files: {list(exported_files.keys())}")

    except Exception as e:
        logger.error(f"Error exporting complete review: {e}")

    return exported_files
